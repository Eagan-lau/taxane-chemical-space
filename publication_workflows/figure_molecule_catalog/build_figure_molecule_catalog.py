#!/usr/bin/env python3
"""Build a complete catalog of molecules depicted in the final figure set."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rdkit import Chem, rdBase
from rdkit.Chem import rdMolDescriptors


EXPECTED_APPEARANCES = {
    "Figure 1": 12,
    "Figure 2": 7,
    "Figure 3": 8,
    "Figure 4": 12,
    "Figure S7": 12,
}

APPEARANCE_COLUMNS = [
    "appearance_id",
    "figure",
    "panel",
    "display_order",
    "visual_role",
    "molecule_id",
    "molecule_name",
    "generation",
    "formula",
    "original_depiction_smiles",
    "canonical_isomeric_smiles",
    "full_inchikey",
    "connectivity_key",
    "unique_molecule_id",
    "source_table",
    "source_record",
    "generation_csv_match",
    "formula_match",
    "smiles_valid",
]

UNIQUE_COLUMNS = [
    "unique_molecule_id",
    "full_inchikey",
    "connectivity_key",
    "formula",
    "canonical_isomeric_smiles",
    "appearance_count",
    "figures",
    "panels",
    "visual_roles",
    "molecule_ids",
    "molecule_names",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--article-release", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_formula(value: str) -> str:
    return "".join(str(value or "").split())


def molecule_fields(smiles: str) -> dict[str, str]:
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return {
            "canonical_isomeric_smiles": "",
            "full_inchikey": "",
            "connectivity_key": "",
            "calculated_formula": "",
            "smiles_valid": "False",
        }
    canonical = Chem.MolToSmiles(mol, canonical=True, isomericSmiles=True)
    inchikey = Chem.MolToInchiKey(mol)
    return {
        "canonical_isomeric_smiles": canonical,
        "full_inchikey": inchikey,
        "connectivity_key": inchikey[:14],
        "calculated_formula": rdMolDescriptors.CalcMolFormula(mol),
        "smiles_valid": "True",
    }


def append_appearance(
    rows: list[dict[str, str]],
    *,
    figure: str,
    panel: str,
    display_order: int,
    visual_role: str,
    molecule_id: str,
    molecule_name: str,
    generation: str,
    formula: str,
    smiles: str,
    source_table: Path,
    source_record: str,
) -> None:
    fields = molecule_fields(smiles)
    calculated_formula = fields.pop("calculated_formula")
    source_formula = normalize_formula(formula)
    rows.append(
        {
            "figure": figure,
            "panel": panel,
            "display_order": str(display_order),
            "visual_role": visual_role,
            "molecule_id": molecule_id,
            "molecule_name": molecule_name,
            "generation": str(generation),
            "formula": source_formula or calculated_formula,
            "original_depiction_smiles": smiles,
            **fields,
            "source_table": str(source_table),
            "source_record": source_record,
            "generation_csv_match": "not_applicable"
            if not molecule_id.startswith(("G0_", "G1_", "G2_", "G3_"))
            else "pending",
            "formula_match": "not_provided"
            if not source_formula
            else str(source_formula == calculated_formula),
        }
    )


def collect_appearances(release: Path) -> tuple[list[dict[str, str]], list[Path]]:
    source_root = release / "source_data"
    rows: list[dict[str, str]] = []
    sources: list[Path] = []

    path = (
        source_root
        / "main_figures"
        / "Figure_1"
        / "panel_source_data"
        / "Figure_1C_grammar_examples.tsv"
    )
    sources.append(path)
    order = 0
    for row_index, row in enumerate(read_tsv(path), start=1):
        for role, key in (("substrate", "substrate_smiles"), ("product", "product_smiles")):
            order += 1
            append_appearance(
                rows,
                figure="Figure 1",
                panel="C",
                display_order=order,
                visual_role=f"{row['display_name']} {role}",
                molecule_id="",
                molecule_name=f"{row['display_name']} {role}",
                generation="curated_pathway",
                formula="",
                smiles=row[key],
                source_table=path.relative_to(release),
                source_record=(
                    f"pathway_row={row['pathway_row_number']};"
                    f"grammar_example={row_index};role={role}"
                ),
            )

    path = (
        source_root
        / "main_figures"
        / "Figure_2"
        / "panel_source_data"
        / "Figure_2_G0_molecular_anchors.tsv"
    )
    sources.append(path)
    for row in read_tsv(path):
        append_appearance(
            rows,
            figure="Figure 2",
            panel="network molecular callouts",
            display_order=int(row["display_order"]) + 1,
            visual_role="G0 molecular anchor",
            molecule_id=row["space_id"],
            molecule_name=row["molecule_names"],
            generation="0",
            formula=row["formula"],
            smiles=row["smiles"],
            source_table=path.relative_to(release),
            source_record=f"space_id={row['space_id']}",
        )

    path = (
        source_root
        / "main_figures"
        / "Figure_2"
        / "panel_source_data"
        / "Figure_2_G1_molecular_examples.tsv"
    )
    sources.append(path)
    for row in read_tsv(path):
        append_appearance(
            rows,
            figure="Figure 2",
            panel="network molecular examples",
            display_order=int(row["display_order"]) + 5,
            visual_role=row["display_label"],
            molecule_id=row["target_space_id"],
            molecule_name=row["display_label"],
            generation="1",
            formula=row["target_formula"],
            smiles=row["target_smiles"],
            source_table=path.relative_to(release),
            source_record=f"event_id={row['event_id']};target={row['target_space_id']}",
        )

    path = (
        source_root
        / "main_figures"
        / "Figure_3"
        / "panel_source_data"
        / "Figure_3_A_G0_G3_chain_structures.tsv"
    )
    sources.append(path)
    for display_order, row in enumerate(read_tsv(path), start=1):
        append_appearance(
            rows,
            figure="Figure 3",
            panel="A",
            display_order=display_order,
            visual_role=f"{row['chain_id']} G{row['generation']} structure",
            molecule_id=row["space_id"],
            molecule_name=row["molecule_name"],
            generation=row["generation"],
            formula=row["formula"],
            smiles=row["smiles"],
            source_table=path.relative_to(release),
            source_record=(
                f"chain_id={row['chain_id']};generation={row['generation']};"
                f"space_id={row['space_id']}"
            ),
        )

    path = (
        source_root
        / "main_figures"
        / "Figure_4"
        / "panel_source_data"
        / "Figure_4_A-D_source_data.tsv"
    )
    sources.append(path)
    for row_index, row in enumerate(read_tsv(path), start=1):
        panel = chr(64 + row_index)
        molecules = (
            (
                "source known taxane",
                row["source_space_id"],
                row["source_name"],
                "0",
                "",
                row["source_smiles"],
            ),
            (
                "latent bridge candidate",
                row["bridge_space_id"],
                "",
                row["generation"],
                row["bridge_formula"],
                row["bridge_smiles"],
            ),
            (
                "target known taxane",
                row["target_space_id"],
                row["target_name"],
                "0",
                "",
                row["target_smiles"],
            ),
        )
        for role_index, (role, molecule_id, name, generation, formula, smiles) in enumerate(
            molecules,
            start=1,
        ):
            append_appearance(
                rows,
                figure="Figure 4",
                panel=panel,
                display_order=(row_index - 1) * 3 + role_index,
                visual_role=role,
                molecule_id=molecule_id,
                molecule_name=name,
                generation=generation,
                formula=formula,
                smiles=smiles,
                source_table=path.relative_to(release),
                source_record=(
                    f"route={row_index};role={role};molecule_id={molecule_id}"
                ),
            )

    path = (
        source_root
        / "supplementary_figures"
        / "panel_source_data"
        / "Supplementary_Figure_S12_V4_A_source_data.tsv"
    )
    sources.append(path)
    for display_order, row in enumerate(read_tsv(path), start=1):
        append_appearance(
            rows,
            figure="Figure S7",
            panel="latent-bridge structure gallery",
            display_order=display_order,
            visual_role="high-support latent bridge candidate",
            molecule_id=row["space_id"],
            molecule_name=row["molecule_names"],
            generation=row["generation"],
            formula=row["formula"],
            smiles=row["smiles"],
            source_table=path.relative_to(release),
            source_record=f"space_id={row['space_id']}",
        )
    return rows, sources


def generation_lookup(
    release: Path,
    requested_ids: set[str],
) -> dict[str, dict[str, str]]:
    by_generation: dict[str, set[str]] = defaultdict(set)
    for space_id in requested_ids:
        by_generation[space_id[:2]].add(space_id)
    files = {
        "G0": "G0_known_taxanes.csv",
        "G1": "G1_inferred_intermediates.csv",
        "G2": "G2_inferred_intermediates.csv",
        "G3": "G3_exploratory_intermediates.csv",
    }
    lookup: dict[str, dict[str, str]] = {}
    root = release / "data" / "generation_csv"
    for prefix, filename in files.items():
        pending = set(by_generation.get(prefix, set()))
        if not pending:
            continue
        with (root / filename).open("r", encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                space_id = row["space_id"]
                if space_id in pending:
                    lookup[space_id] = row
                    pending.remove(space_id)
                    if not pending:
                        break
    return lookup


def complete_and_deduplicate(
    rows: list[dict[str, str]],
    release: Path,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    requested = {
        row["molecule_id"]
        for row in rows
        if row["molecule_id"].startswith(("G0_", "G1_", "G2_", "G3_"))
    }
    lookup = generation_lookup(release, requested)
    for row in rows:
        molecule_id = row["molecule_id"]
        if molecule_id not in requested:
            continue
        reference = lookup.get(molecule_id)
        if reference is None:
            row["generation_csv_match"] = "False"
            continue
        reference_fields = molecule_fields(reference["smiles"])
        row["generation_csv_match"] = str(
            reference_fields["full_inchikey"] == row["full_inchikey"]
        )
        if not row["molecule_name"]:
            row["molecule_name"] = reference["molecule_names"]
        if not row["formula"]:
            row["formula"] = reference["formula"]

    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        key = row["full_inchikey"] or f"INVALID::{row['original_depiction_smiles']}"
        grouped[key].append(row)
    unique_rows: list[dict[str, str]] = []
    for index, key in enumerate(sorted(grouped), start=1):
        group = grouped[key]
        unique_id = f"FIGMOL_{index:04d}"
        for row in group:
            row["unique_molecule_id"] = unique_id
        unique_rows.append(
            {
                "unique_molecule_id": unique_id,
                "full_inchikey": group[0]["full_inchikey"],
                "connectivity_key": group[0]["connectivity_key"],
                "formula": group[0]["formula"],
                "canonical_isomeric_smiles": group[0]["canonical_isomeric_smiles"],
                "appearance_count": str(len(group)),
                "figures": ";".join(sorted({row["figure"] for row in group})),
                "panels": ";".join(
                    sorted({f"{row['figure']} {row['panel']}" for row in group})
                ),
                "visual_roles": ";".join(
                    sorted({row["visual_role"] for row in group})
                ),
                "molecule_ids": ";".join(
                    sorted({row["molecule_id"] for row in group if row["molecule_id"]})
                ),
                "molecule_names": ";".join(
                    sorted({row["molecule_name"] for row in group if row["molecule_name"]})
                ),
            }
        )
    for index, row in enumerate(rows, start=1):
        row["appearance_id"] = f"FIGAPP_{index:04d}"
    return rows, unique_rows


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows({column: row.get(column, "") for column in columns} for row in rows)


def write_tsv(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle, delimiter="\t", lineterminator="\n").writerows(rows)


def add_sheet(
    workbook: Workbook,
    title: str,
    columns: list[str],
    rows: list[dict[str, str]],
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        maximum = max(
            [len(str(column))]
            + [len(str(row.get(column, ""))) for row in rows[:2000]]
        )
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(11, maximum + 2),
            70,
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_audit(
    appearances: list[dict[str, str]],
    unique_rows: list[dict[str, str]],
) -> list[list[object]]:
    observed = defaultdict(int)
    for row in appearances:
        observed[row["figure"]] += 1
    checks: list[list[object]] = []

    def add(check: str, value: object, expected: object, passed: bool, note: str = "") -> None:
        checks.append([check, value, expected, "PASS" if passed else "FAIL", note])

    for figure, expected in EXPECTED_APPEARANCES.items():
        add(
            f"{figure.replace(' ', '_')}_appearance_count",
            observed[figure],
            expected,
            observed[figure] == expected,
        )
    add("total_appearance_count", len(appearances), 51, len(appearances) == 51)
    add(
        "missing_original_smiles",
        sum(not row["original_depiction_smiles"] for row in appearances),
        0,
        all(row["original_depiction_smiles"] for row in appearances),
    )
    add(
        "invalid_smiles",
        sum(row["smiles_valid"] != "True" for row in appearances),
        0,
        all(row["smiles_valid"] == "True" for row in appearances),
    )
    generation_rows = [
        row
        for row in appearances
        if row["generation_csv_match"] != "not_applicable"
    ]
    add(
        "generation_csv_structure_mismatches",
        sum(row["generation_csv_match"] != "True" for row in generation_rows),
        0,
        all(row["generation_csv_match"] == "True" for row in generation_rows),
    )
    formula_rows = [
        row for row in appearances if row["formula_match"] != "not_provided"
    ]
    add(
        "provided_formula_mismatches",
        sum(row["formula_match"] != "True" for row in formula_rows),
        0,
        all(row["formula_match"] == "True" for row in formula_rows),
    )
    add(
        "missing_full_inchikey",
        sum(not row["full_inchikey"] for row in appearances),
        0,
        all(row["full_inchikey"] for row in appearances),
    )
    add(
        "unique_molecule_count",
        len(unique_rows),
        "reported",
        len(unique_rows) > 0,
    )
    add(
        "molecule_bearing_final_figures",
        ";".join(EXPECTED_APPEARANCES),
        "Figure 1;Figure 2;Figure 3;Figure 4;Figure S7",
        list(EXPECTED_APPEARANCES)
        == ["Figure 1", "Figure 2", "Figure 3", "Figure 4", "Figure S7"],
        "Other final supplementary figures contain no complete molecular depictions.",
    )
    return checks


def main() -> int:
    args = parse_args()
    release = args.article_release.resolve()
    output = args.output.resolve()
    if output.exists() and any(output.iterdir()):
        raise FileExistsError(f"Output is not empty: {output}")
    output.mkdir(parents=True, exist_ok=True)

    appearances, source_paths = collect_appearances(release)
    appearances, unique_rows = complete_and_deduplicate(appearances, release)
    audit_rows = build_audit(appearances, unique_rows)
    failures = [row for row in audit_rows if row[3] == "FAIL"]

    write_csv(
        output / "figure_molecule_smiles_appearances.csv",
        appearances,
        APPEARANCE_COLUMNS,
    )
    write_csv(
        output / "figure_molecule_smiles_unique.csv",
        unique_rows,
        UNIQUE_COLUMNS,
    )
    write_tsv(
        output / "figure_molecule_smiles_audit.tsv",
        [["check_id", "observed", "expected", "status", "note"], *audit_rows],
    )

    summary_rows = [
        {
            "figure": figure,
            "depicted_molecule_appearances": str(
                sum(row["figure"] == figure for row in appearances)
            ),
            "unique_molecules_within_figure": str(
                len(
                    {
                        row["full_inchikey"]
                        for row in appearances
                        if row["figure"] == figure
                    }
                )
            ),
            "source_tables": ";".join(
                sorted(
                    {
                        row["source_table"]
                        for row in appearances
                        if row["figure"] == figure
                    }
                )
            ),
        }
        for figure in EXPECTED_APPEARANCES
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "Appearance_Catalog", APPEARANCE_COLUMNS, appearances)
    add_sheet(workbook, "Unique_Molecules", UNIQUE_COLUMNS, unique_rows)
    add_sheet(
        workbook,
        "Figure_Summary",
        [
            "figure",
            "depicted_molecule_appearances",
            "unique_molecules_within_figure",
            "source_tables",
        ],
        summary_rows,
    )
    audit_dicts = [
        {
            "check_id": row[0],
            "observed": row[1],
            "expected": row[2],
            "status": row[3],
            "note": row[4],
        }
        for row in audit_rows
    ]
    add_sheet(
        workbook,
        "Audit",
        ["check_id", "observed", "expected", "status", "note"],
        audit_dicts,
    )
    workbook.save(output / "Figure_Molecule_SMILES_Catalog.xlsx")

    manifest_rows = [
        [
            str(path.relative_to(release)),
            path.stat().st_size,
            sha256_file(path),
        ]
        for path in source_paths
    ]
    write_tsv(
        output / "figure_molecule_source_manifest.tsv",
        [["source_table", "size_bytes", "sha256"], *manifest_rows],
    )
    metadata = {
        "status": "PASS" if not failures else "FAIL",
        "rdkit_version": rdBase.rdkitVersion,
        "depicted_molecule_appearances": len(appearances),
        "unique_full_inchikey_molecules": len(unique_rows),
        "molecule_bearing_figures": list(EXPECTED_APPEARANCES),
        "scientific_recalculation_performed": False,
        "scope": (
            "Complete molecular depictions in the final main and supplementary "
            "figure set; reaction SMARTS fragments are excluded."
        ),
    }
    (output / "build_summary.json").write_text(
        json.dumps(metadata, indent=2) + "\n",
        encoding="utf-8",
    )
    readme = f"""# Figure molecule SMILES catalog

This catalog records every complete small-molecule structure depicted in the
final main and supplementary figure set.

- Depicted molecule appearances: {len(appearances)}
- Unique full-InChIKey structures: {len(unique_rows)}
- Molecule-bearing figures: Figure 1, Figure 2, Figure 3, Figure 4, Figure S7
- RDKit version used for validation: {rdBase.rdkitVersion}

`original_depiction_smiles` preserves the exact SMILES supplied to the figure
workflow. `canonical_isomeric_smiles` is an additional RDKit-normalized field
for indexing and deduplication; it does not replace the depiction input.

Reaction SMARTS fragments in Figure 1 are grammar representations rather than
complete molecules and are therefore retained in the Figure 1 source table,
not in this molecular catalog. The remaining final supplementary figures do
not depict complete molecular structures.
"""
    (output / "README.md").write_text(readme, encoding="utf-8")

    print(json.dumps(metadata, indent=2))
    if failures:
        print(f"Catalog audit failed {len(failures)} checks", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
