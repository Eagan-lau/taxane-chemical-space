from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import count_data_rows, sha256_file, write_tsv


GENERATION_FILES = {
    "0": "G0_known_taxanes.csv",
    "1": "G1_inferred_intermediates.csv",
    "2": "G2_inferred_intermediates.csv",
    "3": "G3_exploratory_intermediates.csv",
}

EXPECTED_GENERATION_COUNTS = {
    "0": 648,
    "1": 15_801,
    "2": 223_823,
    "3": 2_362_766,
}

TABLE_GROUP_TITLES = {
    1: "Reaction provenance, rule-build attrition, evidence tiers, and pathway calibration",
    2: "Primary T1 grammar and G0 activation",
    3: "Generation-resolved chemical-space expansion",
    4: "Authoritative molecular, derivation, application, and rejection data objects",
    5: "Quality-control comparison, robustness, and computational performance",
    6: "Physicochemical descriptors and nearest-G0 similarity",
    7: "Structure-derived functional-state transitions",
    8: "Grammar-use concentration, reaction-edit locality, and elemental deltas",
    9: "Convergence and route multiplicity",
    10: "All latent bridge candidates",
    11: "All directed G0-pair bridge records",
}


def export_generation_csvs(
    nodes_tsv: Path,
    output_dir: Path,
) -> list[dict[str, object]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    handles: dict[str, object] = {}
    writers: dict[str, csv.DictWriter] = {}
    counts = {generation: 0 for generation in GENERATION_FILES}
    try:
        with nodes_tsv.open("r", encoding="utf-8", newline="") as source:
            reader = csv.DictReader(source, delimiter="\t")
            if reader.fieldnames is None or "generation_first" not in reader.fieldnames:
                raise ValueError(f"Missing generation_first in {nodes_tsv}")
            for generation, filename in GENERATION_FILES.items():
                handle = (output_dir / filename).open(
                    "w",
                    encoding="utf-8",
                    newline="",
                )
                handles[generation] = handle
                writer = csv.DictWriter(
                    handle,
                    fieldnames=reader.fieldnames,
                    extrasaction="raise",
                    lineterminator="\n",
                )
                writer.writeheader()
                writers[generation] = writer
            for row in reader:
                generation = row["generation_first"]
                if generation not in writers:
                    raise ValueError(f"Unexpected generation {generation!r}")
                writers[generation].writerow(row)
                counts[generation] += 1
    finally:
        for handle in handles.values():
            handle.close()

    audit: list[dict[str, object]] = []
    for generation, filename in GENERATION_FILES.items():
        path = output_dir / filename
        observed = counts[generation]
        expected = EXPECTED_GENERATION_COUNTS[generation]
        if observed != expected:
            raise ValueError(
                f"{filename}: expected {expected:,} rows, observed {observed:,}"
            )
        audit.append(
            {
                "generation": f"G{generation}",
                "interpretation": (
                    "known taxane seed space"
                    if generation == "0"
                    else "primary near-seed inferred intermediates"
                    if generation in {"1", "2"}
                    else "exploratory inferred intermediates"
                ),
                "file": filename,
                "record_count": observed,
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "contains_smiles": True,
            }
        )
    write_tsv(
        output_dir / "generation_csv_audit.tsv",
        [
            [
                "generation",
                "interpretation",
                "file",
                "record_count",
                "size_bytes",
                "sha256",
                "contains_smiles",
            ],
            *[
                [
                    row["generation"],
                    row["interpretation"],
                    row["file"],
                    row["record_count"],
                    row["size_bytes"],
                    row["sha256"],
                    row["contains_smiles"],
                ]
                for row in audit
            ],
        ],
    )
    return audit


def export_bridge_tables(analysis_dir: Path, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    source_candidates = analysis_dir / "latent_bridge_candidates.tsv"
    bridge_candidates = output_dir / "Table_S10_latent_bridge_candidates.tsv"
    with source_candidates.open("r", encoding="utf-8", newline="") as source:
        reader = csv.DictReader(source, delimiter="\t")
        if reader.fieldnames is None:
            raise ValueError(f"Missing header: {source_candidates}")
        with bridge_candidates.open("w", encoding="utf-8", newline="") as target:
            writer = csv.DictWriter(
                target,
                fieldnames=reader.fieldnames,
                delimiter="\t",
                lineterminator="\n",
            )
            writer.writeheader()
            count = 0
            for row in reader:
                if row.get("latent_bridge_candidate", "").strip().lower() == "true":
                    writer.writerow(row)
                    count += 1
    if count != 227:
        raise ValueError(f"Expected 227 bridge candidates, observed {count}")

    directed_pairs = output_dir / "Table_S11_directed_G0_pair_bridge_support.tsv"
    source_pairs = analysis_dir / "known_G0_pair_bridge_summary.tsv"
    directed_pairs.write_bytes(source_pairs.read_bytes())
    if count_data_rows(directed_pairs) != 309:
        raise ValueError("Directed G0 pair table does not contain 309 records")
    return bridge_candidates, directed_pairs


def _excel_safe(value: str) -> str:
    if len(value) <= 32_767:
        return value
    raise ValueError(
        "A supplementary-table cell exceeds Excel's 32,767-character limit; "
        "the TSV remains authoritative and must not be silently truncated."
    )


def _append_group_to_sheet(
    workbook: Workbook,
    sheet_name: str,
    components: list[dict[str, object]],
) -> tuple[int, list[dict[str, object]]]:
    sheet = workbook.create_sheet(sheet_name)
    total_records = 0
    component_audit: list[dict[str, object]] = []
    current_row = 1
    maximum_columns = 1
    for component in components:
        path = Path(component["path"])
        label = str(component["label"])
        title = str(component["title"])
        section_cell = sheet.cell(
            row=current_row,
            column=1,
            value=f"{label}. {title}",
        )
        section_cell.font = Font(bold=True, color="FFFFFF", size=11)
        section_cell.fill = PatternFill("solid", fgColor="355C7D")
        current_row += 1

        records = 0
        header_row = current_row
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle, delimiter="\t")
            for row_number, row in enumerate(reader, start=1):
                maximum_columns = max(maximum_columns, len(row))
                for column_number, value in enumerate(row, start=1):
                    sheet.cell(
                        row=current_row,
                        column=column_number,
                        value=_excel_safe(value),
                    )
                if row_number > 1:
                    records += 1
                current_row += 1
        for cell in sheet[header_row]:
            if cell.column > maximum_columns:
                break
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="6B8196")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        component_audit.append(
            {
                "component": label,
                "title": title,
                "records": records,
                "authoritative_tsv": path.name,
                "sha256": sha256_file(path),
            }
        )
        total_records += records
        current_row += 2

    sheet.freeze_panes = "A3"
    for column_number in range(1, maximum_columns + 1):
        values = [
            len(str(sheet.cell(row=row, column=column_number).value or ""))
            for row in range(1, min(sheet.max_row, 300) + 1)
        ]
        sheet.column_dimensions[get_column_letter(column_number)].width = min(
            max(max(values, default=8) + 2, 10),
            55,
        )
    return total_records, component_audit


def _style_simple_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="355C7D")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in sheet.iter_cols():
        maximum = max(
            len(str(cell.value or ""))
            for cell in column_cells[: min(len(column_cells), 250)]
        )
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(maximum + 2, 12),
            70,
        )


def build_supplementary_workbook(
    table_groups: dict[int, list[dict[str, object]]],
    workbook_path: Path,
    generation_audit: list[dict[str, object]],
    main_figure_audits: list[Path],
) -> list[dict[str, object]]:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    index = workbook.active
    index.title = "Table_Index"
    index.append(
        [
            "table",
            "title",
            "component_count",
            "records",
            "authoritative_tsvs",
            "sha256",
        ]
    )
    table_audit: list[dict[str, object]] = []

    for number in range(1, 12):
        components = table_groups[number]
        records, component_audit = _append_group_to_sheet(
            workbook,
            f"Table_S{number}",
            components,
        )
        entry = {
            "table": f"Table S{number}",
            "title": TABLE_GROUP_TITLES[number],
            "component_count": len(components),
            "records": records,
            "authoritative_tsvs": ";".join(
                str(component["authoritative_tsv"]) for component in component_audit
            ),
            "sha256": ";".join(
                str(component["sha256"]) for component in component_audit
            ),
            "components": component_audit,
        }
        table_audit.append(entry)
        index.append(
            [
                entry["table"],
                entry["title"],
                entry["component_count"],
                entry["records"],
                entry["authoritative_tsvs"],
                entry["sha256"],
            ]
        )

    data_files = workbook.create_sheet("Generation_CSVs")
    data_files.append(
        [
            "generation",
            "interpretation",
            "file",
            "record_count",
            "size_bytes",
            "sha256",
            "contains_smiles",
        ]
    )
    for row in generation_audit:
        data_files.append(list(row.values()))

    audit_sheet = workbook.create_sheet("Main_Figure_Audits")
    audit_sheet.append(["figure_audit_file", "row_number", "values"])
    for path in main_figure_audits:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.reader(handle, delimiter="\t")
            header = next(reader)
            for row_number, row in enumerate(reader, start=2):
                audit_sheet.append(
                    [
                        path.name,
                        row_number,
                        " | ".join(
                            f"{key}={value}" for key, value in zip(header, row)
                        ),
                    ]
                )

    for sheet in (index, data_files, audit_sheet):
        _style_simple_sheet(sheet)

    workbook.save(workbook_path)
    verified = load_workbook(workbook_path, read_only=True, data_only=True)
    expected_sheets = {
        "Table_Index",
        "Generation_CSVs",
        "Main_Figure_Audits",
        *(f"Table_S{number}" for number in range(1, 12)),
    }
    if set(verified.sheetnames) != expected_sheets:
        raise ValueError("Supplementary workbook sheet inventory failed")
    verified.close()
    return table_audit
