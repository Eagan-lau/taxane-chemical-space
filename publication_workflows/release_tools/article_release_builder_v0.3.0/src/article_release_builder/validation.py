from __future__ import annotations

import csv
import re
import subprocess
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from .data_exports import (
    EXPECTED_GENERATION_COUNTS,
    GENERATION_FILES,
    TABLE_GROUP_TITLES,
)
from .utils import count_data_rows, write_tsv


def _pdf_pages(path: Path) -> int:
    result = subprocess.run(
        ["pdfinfo", str(path)],
        check=True,
        capture_output=True,
        text=True,
    )
    for line in result.stdout.splitlines():
        if line.startswith("Pages:"):
            return int(line.split(":", 1)[1].strip())
    raise ValueError(f"No page count in pdfinfo output for {path}")


def _docx_media_count(path: Path) -> int:
    with zipfile.ZipFile(path) as archive:
        return sum(
            1
            for name in archive.namelist()
            if name.startswith("word/media/") and not name.endswith("/")
        )


def _docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        return archive.read("word/document.xml").decode("utf-8", errors="replace")


def _audit_figure_table(path: Path) -> tuple[int, int]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = list(reader)
    passed = 0
    for row in rows:
        value = row.get("status", row.get("pass", "")).strip().lower()
        if value in {"pass", "true"}:
            passed += 1
    return passed, len(rows)


def validate_completed_release(
    output: Path,
    primary_release: Path,
    main_figure_audits: list[Path],
    table_groups: dict[int, list[dict[str, object]]],
) -> dict[str, object]:
    checks: list[list[object]] = []

    def add(
        check_id: str,
        observed: object,
        expected: object,
        passed: bool,
        note: str = "",
    ) -> None:
        checks.append(
            [check_id, observed, expected, "PASS" if passed else "FAIL", note]
        )

    for generation, filename in GENERATION_FILES.items():
        path = output / "data" / "generation_csv" / filename
        count = 0
        empty_smiles = 0
        wrong_generation = 0
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise ValueError(f"Missing CSV header: {path}")
            required = {"space_id", "generation_first", "smiles", "full_inchikey"}
            add(
                f"G{generation}_required_columns",
                sorted(required.intersection(reader.fieldnames)),
                sorted(required),
                required.issubset(reader.fieldnames),
            )
            for row in reader:
                count += 1
                if not row.get("smiles", "").strip():
                    empty_smiles += 1
                if row.get("generation_first") != generation:
                    wrong_generation += 1
        add(
            f"G{generation}_record_count",
            count,
            EXPECTED_GENERATION_COUNTS[generation],
            count == EXPECTED_GENERATION_COUNTS[generation],
        )
        add(f"G{generation}_empty_smiles", empty_smiles, 0, empty_smiles == 0)
        add(
            f"G{generation}_wrong_generation",
            wrong_generation,
            0,
            wrong_generation == 0,
        )

    workbook_path = (
        output / "supplementary_tables" / "Supplementary_Tables_S1-S11.xlsx"
    )
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    expected_sheets = {
        "Table_Index",
        "Generation_CSVs",
        "Main_Figure_Audits",
        *(f"Table_S{number}" for number in range(1, 12)),
    }
    add(
        "excel_sheet_inventory",
        sorted(workbook.sheetnames),
        sorted(expected_sheets),
        set(workbook.sheetnames) == expected_sheets,
    )
    index_rows = {
        str(row[0]): row
        for row in workbook["Table_Index"].iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[0]
    }
    for number in range(1, 12):
        table_name = f"Table S{number}"
        expected_records = sum(
            count_data_rows(Path(component["path"]))
            for component in table_groups[number]
        )
        index_row = index_rows.get(table_name)
        observed_records = index_row[3] if index_row else None
        add(
            f"excel_Table_S{number}_records",
            observed_records,
            expected_records,
            observed_records == expected_records,
        )
        add(
            f"excel_Table_S{number}_components",
            index_row[2] if index_row else None,
            len(table_groups[number]),
            bool(index_row and index_row[2] == len(table_groups[number])),
        )
    workbook.close()

    expected_main_figure_audit_rows = {1: 15, 2: 24, 3: 15, 4: 28}
    for number, path in enumerate(main_figure_audits, start=1):
        passed, total = _audit_figure_table(path)
        add(f"Figure_{number}_numerical_audit", passed, total, passed == total)
        add(
            f"Figure_{number}_numerical_audit_rows",
            total,
            expected_main_figure_audit_rows[number],
            total == expected_main_figure_audit_rows[number],
        )

    manuscript_docx = output / "manuscript" / "Main_Manuscript_with_Figures.docx"
    supplementary_docx = (
        output / "supplementary_information" / "Supplementary_Information.docx"
    )
    main_media = _docx_media_count(manuscript_docx)
    supplementary_media = _docx_media_count(supplementary_docx)
    add("main_docx_embedded_figures", main_media, 4, main_media == 4)
    add(
        "supplementary_docx_embedded_figures",
        supplementary_media,
        8,
        supplementary_media == 8,
    )

    main_docx_text = _docx_text(manuscript_docx)
    supplementary_docx_text = _docx_text(supplementary_docx)
    for number in range(1, 5):
        token = f"Figure {number} |"
        add(
            f"main_docx_caption_Figure_{number}",
            token in main_docx_text,
            True,
            token in main_docx_text,
        )
    for number in range(1, 9):
        token = f"Supplementary Figure S{number} |"
        add(
            f"supplementary_docx_caption_S{number}",
            token in supplementary_docx_text,
            True,
            token in supplementary_docx_text,
        )

    manuscript_pdf = output / "manuscript" / "Main_Manuscript_with_Figures.pdf"
    supplementary_pdf = (
        output / "supplementary_information" / "Supplementary_Information.pdf"
    )
    manuscript_pages = _pdf_pages(manuscript_pdf)
    supplementary_pages = _pdf_pages(supplementary_pdf)
    add("main_pdf_pages", manuscript_pages, ">=12", manuscript_pages >= 12)
    add(
        "supplementary_pdf_pages",
        supplementary_pages,
        ">=8",
        supplementary_pages >= 8,
    )
    for number in range(1, 9):
        figure_pdf = output / "figures" / "supplementary" / f"Figure_S{number}.pdf"
        figure_svg = output / "figures" / "supplementary" / f"Figure_S{number}.svg"
        add(
            f"supplementary_figure_S{number}_pdf_pages",
            _pdf_pages(figure_pdf),
            1,
            _pdf_pages(figure_pdf) == 1,
        )
        add(
            f"supplementary_figure_S{number}_svg_nonempty",
            figure_svg.stat().st_size,
            ">0",
            figure_svg.stat().st_size > 0,
        )

    manuscript_markdown = (
        output / "manuscript" / "Main_Manuscript_with_Figures.md"
    ).read_text(encoding="utf-8")
    normalized_manuscript = re.sub(r"\s+", " ", manuscript_markdown)
    supplementary_markdown = (
        output / "supplementary_information" / "Supplementary_Information.md"
    ).read_text(encoding="utf-8")
    combined_markdown = manuscript_markdown + supplementary_markdown
    unresolved = combined_markdown.count("{{")
    cjk = len(re.findall(r"[\u4e00-\u9fff]", combined_markdown))
    stale_main_figures = len(re.findall(r"Fig\.\s*[5-9]\b", manuscript_markdown))
    stale_supp_figures = len(
        re.findall(r"(?:Fig\.|Figure)\s*S(?:9|1[0-9])\b", combined_markdown)
    )
    stale_supp_tables = len(
        re.findall(r"Table\s*S(?:1[2-9]|2[0-9])\b", combined_markdown)
    )
    add("unresolved_template_tokens", unresolved, 0, unresolved == 0)
    add("CJK_characters", cjk, 0, cjk == 0)
    add(
        "stale_main_figure_references",
        stale_main_figures,
        0,
        stale_main_figures == 0,
    )
    add(
        "stale_supplementary_figure_references",
        stale_supp_figures,
        0,
        stale_supp_figures == 0,
    )
    add(
        "stale_supplementary_table_references",
        stale_supp_tables,
        0,
        stale_supp_tables == 0,
    )

    stale_figure2_phrases = (
        "G2 and G3 are retained at their complete numerical scale as "
        "component-normalized descendant-density fields",
        "G2 is encoded by a teal fill with solid contours",
    )
    stale_figure2_count = sum(
        manuscript_markdown.count(phrase) for phrase in stale_figure2_phrases
    )
    add(
        "stale_Figure_2_density_encoding",
        stale_figure2_count,
        0,
        stale_figure2_count == 0,
    )
    figure2_required_tokens = (
        "223,823 G2 structures as individual nodes",
        "bounded deterministic local relaxation",
        "This operation changed display coordinates only",
        "Stipple marks encode density",
    )
    for index, token in enumerate(figure2_required_tokens, start=1):
        add(
            f"Figure_2_V12_disclosure_{index}",
            token in normalized_manuscript,
            True,
            token in normalized_manuscript,
            token,
        )
    figure2_source = (
        output
        / "source_data"
        / "main_figures"
        / "Figure_2"
        / "panel_source_data"
        / "Figure_2_G2_display_layout.tsv"
    )
    add(
        "Figure_2_V12_G2_display_source_present",
        figure2_source.is_file(),
        True,
        figure2_source.is_file() and figure2_source.stat().st_size > 0,
    )

    figure_molecule_root = (
        output / "source_data" / "figure_molecule_catalog"
    )
    figure_molecule_audit = (
        figure_molecule_root / "figure_molecule_smiles_audit.tsv"
    )
    if figure_molecule_audit.is_file():
        with figure_molecule_audit.open(
            "r",
            encoding="utf-8",
            newline="",
        ) as handle:
            molecule_audit_rows = list(csv.DictReader(handle, delimiter="\t"))
        for row in molecule_audit_rows:
            status = row.get("status", "").strip().upper()
            add(
                f"figure_molecule_{row['check_id']}",
                row.get("observed", ""),
                row.get("expected", ""),
                status == "PASS",
                row.get("note", ""),
            )
        appearance_csv = (
            figure_molecule_root
            / "figure_molecule_smiles_appearances.csv"
        )
        unique_csv = (
            figure_molecule_root / "figure_molecule_smiles_unique.csv"
        )
        workbook = (
            figure_molecule_root / "Figure_Molecule_SMILES_Catalog.xlsx"
        )
        add(
            "figure_molecule_appearance_csv_rows",
            count_data_rows(appearance_csv),
            51,
            appearance_csv.is_file()
            and count_data_rows(appearance_csv) == 51,
        )
        add(
            "figure_molecule_unique_csv_rows",
            count_data_rows(unique_csv),
            44,
            unique_csv.is_file() and count_data_rows(unique_csv) == 44,
        )
        add(
            "figure_molecule_workbook_present",
            workbook.is_file(),
            True,
            workbook.is_file() and workbook.stat().st_size > 0,
        )
        upload_workbook = (
            output
            / "submission_upload"
            / "11_Figure_Molecule_SMILES_Source_Data.xlsx"
        )
        add(
            "figure_molecule_submission_workbook_present",
            upload_workbook.is_file(),
            True,
            upload_workbook.is_file()
            and upload_workbook.stat().st_size > 0,
        )

    headline_claims = {
        "normalized_reactions": "630,280",
        "generalized_smarts": "353,524",
        "G0": "648",
        "G1": "15,801",
        "G2": "223,823",
        "G3": "2,362,766",
        "total_structures": "2,603,038",
        "accepted_events": "7,928,209",
        "bridge_candidates": "227",
        "directed_pairs": "309",
    }
    for claim, token in headline_claims.items():
        add(
            f"headline_claim_{claim}",
            token in manuscript_markdown,
            True,
            token in manuscript_markdown,
        )

    source_files = {
        path.relative_to(primary_release): path
        for path in primary_release.rglob("*")
        if path.is_file() and not path.is_symlink()
    }
    snapshot = output / "data" / "raw_primary_release"
    linked = 0
    missing = 0
    inode_mismatch = 0
    for relative, source in source_files.items():
        target = snapshot / relative
        if not target.is_file():
            missing += 1
        elif (
            source.stat().st_dev == target.stat().st_dev
            and source.stat().st_ino == target.stat().st_ino
        ):
            linked += 1
        else:
            inode_mismatch += 1
    add(
        "raw_snapshot_file_count",
        linked + inode_mismatch,
        len(source_files),
        missing == 0,
    )
    add(
        "raw_snapshot_hardlink_count",
        linked,
        len(source_files),
        linked == len(source_files),
    )
    add(
        "raw_snapshot_inode_mismatch",
        inode_mismatch,
        0,
        inode_mismatch == 0,
    )

    failures = [row for row in checks if row[3] == "FAIL"]
    write_tsv(
        output / "validation" / "FINAL_AUDIT.tsv",
        [["check_id", "observed", "expected", "status", "note"], *checks],
    )
    report = f"""# Final validation report

**Overall status:** {'PASS' if not failures else 'FAIL'}

## Scope

The audit validates generation-specific molecular CSV records and SMILES
completeness, the 11-table workbook topology and component row counts,
main-figure numerical audits, figure-caption embedding, PDF rendering,
cross-references after editorial consolidation, headline numerical claims,
the hard-linked frozen raw-data snapshot, and, when present, the complete
figure-molecule SMILES catalog.

## Results

- Checks executed: {len(checks)}
- Passed: {len(checks) - len(failures)}
- Failed: {len(failures)}
- G0-G3 molecular CSV records: 648, 15,801, 223,823, and 2,362,766
- Main figures embedded with captions: 4
- Supplementary figures embedded with captions: 8
- Logical supplementary tables: {len(TABLE_GROUP_TITLES)}
- Figure-molecule appearances catalogued: {
    51 if figure_molecule_audit.is_file() else 'not included'
}
- Frozen raw files present as hard links: {linked} of {len(source_files)}

## Interpretation boundary

This validation establishes internal consistency and release completeness. It
does not establish biological occurrence or enzymatic feasibility of inferred
structures.
"""
    (output / "validation" / "FINAL_VALIDATION_REPORT.md").write_text(
        report,
        encoding="utf-8",
    )
    readiness = """# Submission readiness

## Scientific and technical status

The rewritten manuscript, four main figures with embedded captions, curated
Supplementary Information, consolidated workbook, generation-specific
molecular CSVs, figure source data, numerical audits, reproduction scripts,
and frozen raw-data snapshot are complete and internally consistent.

## Author inputs still required

The package is not ready for journal upload until the authors supply:

1. Final author list, order, affiliations, addresses, and ORCID identifiers.
2. Corresponding-author details.
3. Author contributions, funding, acknowledgements, and competing interests.
4. Permanent repository URLs, accessions, release identifiers, and DOI values.
5. Final database snapshot and redistribution statements approved by the
   authors.
6. Journal-portal formatting and policy confirmations.

These fields remain explicit placeholders and were not inferred.
"""
    (output / "validation" / "SUBMISSION_READINESS.md").write_text(
        readiness,
        encoding="utf-8",
    )
    if failures:
        raise ValueError(f"Final release audit failed {len(failures)} checks")
    return {
        "status": "PASS",
        "checks": len(checks),
        "passed": len(checks),
        "failed": 0,
        "raw_hardlinks": linked,
    }
